import tkinter as tk
from tkinter import messagebox, filedialog
from tkinter import ttk
from ttkthemes import ThemedStyle
import smtplib
from email.message import EmailMessage
import openpyxl


# Función para enviar correo electrónico con formato HTML
def enviar_correos():
    try:
        # Configuración del servidor SMTP y credenciales
        smtp_server = smtp_server_entry.get()
        smtp_port = int(smtp_port_entry.get())
        sender_email = sender_email_entry.get()
        password = password_entry.get()

        # Seleccionar el archivo Excel manualmente
        ruta_archivo = filedialog.askopenfilename(title="Seleccione el archivo Excel", filetypes=(
        ("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")))
        if not ruta_archivo:
            messagebox.showwarning("Advertencia", "Debe seleccionar un archivo Excel.")
            return

        # Leer correos electrónicos desde el archivo Excel
        workbook = openpyxl.load_workbook(ruta_archivo)
        sheet = workbook.active
        total_correos = sheet.max_row
        correos_enviados = 0
        errores = 0

        # Configurar la barra de progreso
        progreso.config(maximum=total_correos)

        # Enviar correos electrónicos
        for row in sheet.iter_rows(min_row=2, values_only=True):
            recipient_email = row[0]  # La dirección de correo electrónico se encuentra en la primera columna (comuna A)
            message = EmailMessage()
            message.add_alternative(contenido_entry.get("1.0", tk.END), subtype='html')
            message["Subject"] = asunto_entry.get()
            message["From"] = sender_email
            message["To"] = recipient_email

            try:
                # Establecer conexión con el servidor SMTP y enviar el correo electrónico
                server = smtplib.SMTP_SSL(smtp_server, smtp_port)
                server.login(sender_email, password)
                server.send_message(message)
                server.quit()

                correos_enviados += 1
            except Exception as e:
                print(f"Error al enviar el correo electrónico a {recipient_email}: {str(e)}")
                errores += 1

            # Actualizar la barra de progreso
            progreso.step(1)
            ventana.update_idletasks()

        messagebox.showinfo("Resultados", f"Correos enviados: {correos_enviados}\nErrores: {errores}")

    except Exception as e:
        messagebox.showerror("Error", f"Error al enviar los correos electrónicos: {str(e)}")


# Crear la ventana principal
ventana = tk.Tk()
ventana.title("By IBZAN - Enviar Correos Electrónicos")
ventana.geometry("400x800")  # Establecer tamaño de la ventana

# Establecer el icono de la ventana
ventana.iconbitmap(r"C:\\Users\\zero\\Documents\\icono.ico")

# Estilo uniforme para todos los widgets
style = ThemedStyle(ventana)
style.set_theme("plastik")

# Crear etiquetas y campos de entrada con estilo moderno
smtp_server_label = ttk.Label(ventana, text="Servidor SMTP:")
smtp_server_label.pack()
smtp_server_entry = ttk.Entry(ventana, width=40)
smtp_server_entry.pack()

smtp_port_label = ttk.Label(ventana, text="Puerto SMTP:")
smtp_port_label.pack()
smtp_port_entry = ttk.Entry(ventana, width=40)
smtp_port_entry.pack()

sender_email_label = ttk.Label(ventana, text="Correo Electrónico (Remitente):")
sender_email_label.pack()
sender_email_entry = ttk.Entry(ventana, width=40)
sender_email_entry.pack()

password_label = ttk.Label(ventana, text="Contraseña:")
password_label.pack()
password_entry = ttk.Entry(ventana, show="*", width=40)  # Mostrar asteriscos para ocultar la contraseña
password_entry.pack()

asunto_label = ttk.Label(ventana, text="Asunto:")
asunto_label.pack()
asunto_entry = ttk.Entry(ventana, width=40)
asunto_entry.pack()

contenido_label = ttk.Label(ventana, text="Contenido (HTML):")
contenido_label.pack()
contenido_entry = tk.Text(ventana, height=5, width=40)
contenido_entry.pack()

# Botón de seleccionar archivo Excel con estilo moderno
seleccionar_archivo_button = ttk.Button(ventana, text="Seleccionar Archivo Excel", command=enviar_correos)
seleccionar_archivo_button.pack(pady=10)

# Barra de progreso más delgada y atractiva
style.configure("TProgressbar", thickness=30, troughcolor="#333333", troughrelief="flat", background="#00FF00")
progreso = ttk.Progressbar(ventana, orient="horizontal", length=300, mode="determinate", style="TProgressbar")
progreso.pack(pady=10)

# Línea divisoria
ttk.Separator(ventana, orient="horizontal").pack(fill="x", pady=10)

# Texto
texto = tk.Label(ventana, text="ULTRAHOST")
texto.pack(pady=10)

# Texto
texto = tk.Label(ventana, text="Diseño de páginas web, diseño de aplicaciones,")
texto.pack(pady=10)

# Texto
texto = tk.Label(ventana, text="hosting y dominios. Contáctenos.")
texto.pack(pady=10)

# Botón de página web
pagina_web_button = tk.Button(ventana, text="PAGINA WEB", bg="#4CAF50", fg="white", cursor="hand2", command=lambda: web_redirect("https://www.ultrahost.uk"), width=40)
pagina_web_button.pack(pady=10)

# Botón de WhatsApp
whatsapp_button = tk.Button(ventana, text="WHATSAPP", bg="#4CAF50", fg="white", cursor="hand2", command=lambda: web_redirect("https://wa.me/447418353168"), width=40)
whatsapp_button.pack(pady=10)

def web_redirect(url):
    import webbrowser
    webbrowser.open_new(url)

# Iniciar el bucle principal de la aplicación
ventana.mainloop()
